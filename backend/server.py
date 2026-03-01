from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import random
import string
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from fastapi.responses import StreamingResponse
import base64

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Settings
JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24

# Master OTP Email
MASTER_EMAIL = os.environ.get('MASTER_EMAIL', 'admin@edudham.com')

# Create the main app without public docs in production
ENV = os.environ.get('ENV', 'production')
app = FastAPI(
    docs_url="/docs" if ENV == 'development' else None,
    redoc_url="/redoc" if ENV == 'development' else None,
    openapi_url="/openapi.json" if ENV == 'development' else None,
)
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# ============ MODELS ============

class Category(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class CategoryCreate(BaseModel):
    name: str

class UserRole(BaseModel):
    role: str  # 'admin', 'manager', 'student'
    university_id: Optional[str] = None  # Only for managers

class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    password_hash: str
    name: str
    role: str  # 'admin', 'manager', 'student'
    university_id: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: Optional[str] = 'student'
    university_id: Optional[str] = None

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: str
    university_id: Optional[str] = None

class OTPRequest(BaseModel):
    email: EmailStr

class OTPVerify(BaseModel):
    email: EmailStr
    otp: str
    new_password: str

class OTP(BaseModel):
    model_config = ConfigDict(extra="ignore")
    email: str
    otp: str
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    expires_at: str

class HomepageConfig(BaseModel):
    hero_title: str = "Find Your Perfect"
    hero_title_highlight: str = "College Match"
    hero_subtitle: str = "Explore universities across Uttar Pradesh. Compare courses, fees, and facilities to make the right choice for your future."
    cta_text: str = "Search"
    background_images: List[str] = [
        "https://images.unsplash.com/photo-1680084521806-b408d976e3e7?crop=entropy&cs=srgb&fm=jpg&q=85",
        "https://images.unsplash.com/photo-1562774053-701939374585?crop=entropy&cs=srgb&fm=jpg&q=85",
        "https://images.unsplash.com/photo-1541339907198-e08756dedf3f?crop=entropy&cs=srgb&fm=jpg&q=85",
        "https://images.unsplash.com/photo-1498243691581-b145c3f54a5a?crop=entropy&cs=srgb&fm=jpg&q=85",
    ]
    slide_interval_ms: int = 5000
    # Branding
    site_name: str = "Edu Dham"
    logo_url: Optional[str] = ""
    # Text colours
    hero_title_color: str = "#ffffff"
    hero_highlight_color: str = "#f97316"
    hero_subtitle_color: str = "#cbd5e1"
    # Footer
    show_footer: bool = False
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class FeeStructure(BaseModel):
    course: str
    duration: str
    annual_fee: float

class Course(BaseModel):
    course_name: str
    description: str = ""
    duration: str = ""
    fees: float = 0.0
    category: str = ""

class University(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    location: str
    state: Optional[str] = ""
    university_categories: List[str] = []  # e.g., ["Engineering", "Medical"]
    main_photo: Optional[str] = ""  # Now stores base64 or URL or empty for placeholder
    photo_gallery: List[str] = []
    description: str
    courses: List[Course] = []
    placement_percentage: float
    rating: float = 0.0
    tags: List[str] = []
    contact_details: Dict[str, str] = {}
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class UniversityCreate(BaseModel):
    name: str
    location: str
    state: Optional[str] = ""
    university_categories: List[str] = []
    main_photo: Optional[str] = ""
    photo_gallery: List[str] = []
    description: str
    courses: List[Course] = []
    placement_percentage: float
    rating: float = 0.0
    tags: List[str] = []
    contact_details: Dict[str, str] = {}

class UniversityUpdate(BaseModel):
    name: Optional[str] = None
    location: Optional[str] = None
    state: Optional[str] = None
    university_categories: Optional[List[str]] = None
    main_photo: Optional[str] = None
    photo_gallery: Optional[List[str]] = None
    description: Optional[str] = None
    courses: Optional[List[Course]] = None
    placement_percentage: Optional[float] = None
    rating: Optional[float] = None
    tags: Optional[List[str]] = None
    contact_details: Optional[Dict[str, str]] = None

class Application(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    university_id: str
    university_name: str
    name: str
    email: EmailStr
    phone: str
    course_interest: str
    short_note: str
    status: str = 'pending'  # pending, contacted, accepted, rejected
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ApplicationCreate(BaseModel):
    university_id: str
    university_name: str
    name: str
    email: EmailStr
    phone: str
    course_interest: str
    short_note: str

class ApplicationResponse(BaseModel):
    id: str
    university_id: str
    university_name: str
    name: str
    email: str
    phone: str
    course_interest: str
    short_note: str
    status: str
    created_at: str

# ============ UTILITY FUNCTIONS ============

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, email: str, role: str, university_id: Optional[str] = None) -> str:
    payload = {
        'user_id': user_id,
        'email': email,
        'role': role,
        'university_id': university_id,
        'exp': datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_token(token: str) -> Dict:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> Dict:
    token = credentials.credentials
    return decode_token(token)

def generate_otp() -> str:
    return ''.join(random.choices(string.digits, k=6))

# ============ HOMEPAGE CONFIG ENDPOINTS ============

# Default config singleton
DEFAULT_HOMEPAGE_CONFIG = HomepageConfig()

@api_router.get("/homepage-config")
async def get_homepage_config():
    """Public endpoint — returns the current homepage configuration."""
    doc = await db.homepage_config.find_one({"_id": "singleton"}, {"_id": 0})
    if doc:
        return doc
    return DEFAULT_HOMEPAGE_CONFIG.model_dump()

@api_router.put("/homepage-config")
async def update_homepage_config(
    config: HomepageConfig,
    current_user: Dict = Depends(get_current_user)
):
    """Admin-only endpoint — update the homepage configuration."""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    data = config.model_dump()
    data['updated_at'] = datetime.now(timezone.utc).isoformat()
    await db.homepage_config.update_one(
        {"_id": "singleton"},
        {"$set": data},
        upsert=True
    )
    return data

# ============ AUTH ENDPOINTS ============

@api_router.post("/auth/register")
async def register(
    user_data: UserCreate,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(HTTPBearer(auto_error=False))
):
    # Admin and manager roles can only be created by an existing admin
    if user_data.role in ['admin', 'manager']:
        if not credentials:
            raise HTTPException(status_code=403, detail="Only admins can create admin or manager accounts")
        try:
            current_user = decode_token(credentials.credentials)
            if current_user.get('role') != 'admin':
                raise HTTPException(status_code=403, detail="Only admins can create admin or manager accounts")
        except HTTPException:
            raise HTTPException(status_code=403, detail="Only admins can create admin or manager accounts")

    # Force role to 'student' for public self-registration
    safe_role = user_data.role if user_data.role in ['admin', 'manager'] else 'student'

    existing = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")

    user = User(
        email=user_data.email,
        password_hash=hash_password(user_data.password),
        name=user_data.name,
        role=safe_role,
        university_id=user_data.university_id
    )

    await db.users.insert_one(user.model_dump())

    token = create_token(user.id, user.email, user.role, user.university_id)

    return {
        "token": token,
        "user": UserResponse(
            id=user.id,
            email=user.email,
            name=user.name,
            role=user.role,
            university_id=user.university_id
        )
    }

@api_router.post("/auth/login")
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user or not verify_password(credentials.password, user['password_hash']):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_token(user['id'], user['email'], user['role'], user.get('university_id'))
    
    return {
        "token": token,
        "user": UserResponse(
            id=user['id'],
            email=user['email'],
            name=user['name'],
            role=user['role'],
            university_id=user.get('university_id')
        )
    }

@api_router.post("/auth/request-otp")
async def request_otp(request: OTPRequest):
    user = await db.users.find_one({"email": request.email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    otp_code = generate_otp()
    otp = OTP(
        email=request.email,
        otp=otp_code,
        expires_at=(datetime.now(timezone.utc) + timedelta(minutes=10)).isoformat()
    )
    
    await db.otps.delete_many({"email": request.email})
    await db.otps.insert_one(otp.model_dump())
    
    return {
        "message": f"OTP sent to master email {MASTER_EMAIL}",
        "otp": otp_code,
        "master_email": MASTER_EMAIL
    }

@api_router.post("/auth/verify-otp")
async def verify_otp(data: OTPVerify):
    otp_record = await db.otps.find_one({"email": data.email}, {"_id": 0})
    if not otp_record:
        raise HTTPException(status_code=404, detail="OTP not found")
    
    if datetime.fromisoformat(otp_record['expires_at']) < datetime.now(timezone.utc):
        raise HTTPException(status_code=400, detail="OTP expired")
    
    if otp_record['otp'] != data.otp:
        raise HTTPException(status_code=400, detail="Invalid OTP")
    
    new_hash = hash_password(data.new_password)
    await db.users.update_one(
        {"email": data.email},
        {"$set": {"password_hash": new_hash}}
    )
    
    await db.otps.delete_many({"email": data.email})
    
    return {"message": "Password reset successful"}

# ============ UNIVERSITY ENDPOINTS ============

@api_router.post("/universities/upload-photo")
async def upload_university_photo(
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """Upload a university photo and return a base64 data URL."""
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    # Validate file type
    if not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="File must be an image")
    
    # Read the file and convert to base64
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:  # 5MB limit
        raise HTTPException(status_code=400, detail="Image must be less than 5MB")
    
    b64 = base64.b64encode(contents).decode('utf-8')
    data_url = f"data:{file.content_type};base64,{b64}"
    
    return {"photo_url": data_url}

@api_router.post("/universities", response_model=University)
async def create_university(university_data: UniversityCreate, current_user: Dict = Depends(get_current_user)):
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can create universities")
    
    university = University(**university_data.model_dump())
    await db.universities.insert_one(university.model_dump())
    return university

@api_router.get("/universities/filter-options")
async def get_university_filter_options():
    """Return unique locations and categories from the universities collection."""
    all_unis = await db.universities.find({}, {"_id": 0, "location": 1, "university_categories": 1, "university_category": 1}).to_list(1000)
    locations = sorted(set(u['location'].strip() for u in all_unis if u.get('location', '').strip()))
    categories_docs = await db.categories.find({}, {"_id": 0, "name": 1}).to_list(1000)
    dynamic_cats = [c['name'] for c in categories_docs]
    
    # Merge with existing ones for safety, but primary should be dynamic
    cats = set(dynamic_cats)
    for u in all_unis:
        for c in u.get('university_categories', []):
            if c and c.strip():
                cats.add(c.strip())
        legacy = u.get('university_category', '')
        if legacy and legacy.strip():
            cats.add(legacy.strip())
    categories = sorted(cats)
    return {"locations": locations, "categories": categories}

# ============ CATEGORY CRUD ENDPOINTS ============

@api_router.post("/categories", response_model=Category)
async def create_category(cat_data: CategoryCreate, current_user: Dict = Depends(get_current_user)):
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    
    # Check if exists
    existing = await db.categories.find_one({"name": {"$regex": f"^{cat_data.name}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Category already exists")
        
    category = Category(name=cat_data.name)
    await db.categories.insert_one(category.model_dump())
    return category

@api_router.get("/categories", response_model=List[Category])
async def get_categories():
    return await db.categories.find({}, {"_id": 0}).to_list(1000)

@api_router.put("/categories/{category_id}", response_model=Category)
async def update_category(category_id: str, cat_data: CategoryCreate, current_user: Dict = Depends(get_current_user)):
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
        
    result = await db.categories.update_one(
        {"id": category_id},
        {"$set": {"name": cat_data.name}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
        
    return await db.categories.find_one({"id": category_id}, {"_id": 0})

@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str, current_user: Dict = Depends(get_current_user)):
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
        
    result = await db.categories.delete_one({"id": category_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
        
    return {"message": "Category deleted"}

@api_router.get("/universities", response_model=List[University])
async def get_universities(
    search: Optional[str] = None,
    location: Optional[str] = None,
    state: Optional[str] = None,
    course_type: Optional[str] = None,
    category: Optional[str] = None,
    min_fee: Optional[float] = None,
    max_fee: Optional[float] = None,
    min_rating: Optional[float] = None,
    min_placement: Optional[float] = None
):
    query = {}
    
    if search:
        query['$or'] = [
            {'name': {'$regex': search, '$options': 'i'}},
            {'tags': {'$regex': search, '$options': 'i'}},
            {'location': {'$regex': search, '$options': 'i'}},
            {'state': {'$regex': search, '$options': 'i'}},
            {'courses.course_name': {'$regex': search, '$options': 'i'}},
            {'courses.category': {'$regex': search, '$options': 'i'}},
            {'description': {'$regex': search, '$options': 'i'}},
            {'university_categories': {'$elemMatch': {'$regex': search, '$options': 'i'}}},
        ]
    
    if location:
        query['location'] = {'$regex': location, '$options': 'i'}
    
    if state:
        query['state'] = {'$regex': state, '$options': 'i'}
    
    if course_type:
        query['courses.course_name'] = {'$regex': course_type, '$options': 'i'}

    if category:
        # Match if the given category is in the university_categories array, or matches legacy field
        query['$or'] = query.get('$or', []) + [
            {'university_categories': category},
            {'university_category': {'$regex': category, '$options': 'i'}},
        ]
    
    if min_rating is not None:
        query['rating'] = {'$gte': min_rating}
    
    if min_placement is not None:
        query['placement_percentage'] = {'$gte': min_placement}
    
    universities = await db.universities.find(query, {"_id": 0}).to_list(1000)
    
    if min_fee is not None or max_fee is not None:
        filtered = []
        for uni in universities:
            if uni.get('fee_structure'):
                fees = [fs['annual_fee'] for fs in uni['fee_structure']]
                if fees:
                    avg_fee = sum(fees) / len(fees)
                    if min_fee is not None and avg_fee < min_fee:
                        continue
                    if max_fee is not None and avg_fee > max_fee:
                        continue
            filtered.append(uni)
        universities = filtered
    
    return universities

@api_router.get("/universities/bulk-template/download")
async def download_bulk_template(current_user: Dict = Depends(get_current_user)):
    """Download a sample Excel template for bulk university upload."""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Universities"

    # Header styling
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    headers = [
        "name", "location", "state", "description",
        "courses", "placement_percentage", "rating",
        "tags", "email", "phone", "website"
    ]
    
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.row_dimensions[1].height = 22

    # Sample data rows
    sample_rows = [
        [
            "Example University of Technology",
            "Lucknow",
            "Uttar Pradesh",
            "A premier institute offering cutting-edge engineering and technology programs.",
            "B.Tech, M.Tech, MBA, MCA",
            85.5,
            4.2,
            "Engineering, Technology, Research",
            "info@eut.ac.in",
            "+91-9876543210",
            "https://eut.ac.in"
        ],
        [
            "National College of Science",
            "Kanpur",
            "Uttar Pradesh",
            "Leading science college with state-of-the-art laboratories and experienced faculty.",
            "B.Sc, M.Sc, Ph.D",
            78.0,
            3.9,
            "Science, Research, Academia",
            "contact@ncs.ac.in",
            "+91-9123456780",
            "https://ncs.ac.in"
        ]
    ]
    
    data_fill = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
    for row_data in sample_rows:
        ws.append(row_data)
        row_idx = ws.max_row
        for cell in ws[row_idx]:
            cell.fill = data_fill
            cell.alignment = Alignment(wrap_text=True)

    # Adjust column widths
    col_widths = [35, 20, 25, 50, 35, 22, 10, 30, 25, 18, 30]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # Instructions sheet
    ws2 = wb.create_sheet(title="Instructions")
    ws2.column_dimensions['A'].width = 80
    instructions = [
        ["📋 BULK UNIVERSITY UPLOAD - INSTRUCTIONS"],
        [""],
        ["REQUIRED FIELDS (must be filled):"],
        ["  • name              : Full university name"],
        ["  • location          : City / District name"],
        ["  • state             : State name (e.g., Uttar Pradesh, Delhi, Bihar)"],
        ["  • description       : Short description of the university"],
        ["  • placement_percentage : Number between 0 and 100 (e.g., 85.5)"],
        [""],
        ["OPTIONAL FIELDS:"],
        ["  • courses           : Comma-separated course names (e.g., B.Tech, MBA, MCA)"],
        ["  • rating            : Number between 0 and 5 (e.g., 4.2)"],
        ["  • tags              : Comma-separated keywords for search"],
        ["  • email             : Contact email of the university"],
        ["  • phone             : Contact phone number"],
        ["  • website           : University website URL"],
        [""],
        ["NOTES:"],
        ["  • Do NOT change or delete the header row (row 1)."],
        ["  • You can delete the 2 sample rows before uploading."],
        ["  • Photos can be added individually after bulk upload via the Edit button."],
        ["  • Maximum 200 universities per upload."],
    ]
    for row in instructions:
        ws2.append(row)
    ws2['A1'].font = Font(bold=True, size=13, color="1E3A5F")
    for r in range(3, ws2.max_row + 1):
        ws2.cell(row=r, column=1).font = Font(size=10)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=university_bulk_upload_template.xlsx"}
    )

@api_router.post("/universities/bulk-upload")
async def bulk_upload_universities(
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """Bulk upload universities from an Excel file."""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    contents = await file.read()
    try:
        wb = load_workbook(BytesIO(contents))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read Excel file. Please use the provided template.")
    
    # Read headers from first row
    headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
    
    required_fields = {'name', 'location', 'description', 'placement_percentage'}
    missing = required_fields - set(headers)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {', '.join(missing)}. Please use the provided template."
        )
    
    created = []
    errors = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = dict(zip(headers, row))
        
        # Skip empty rows
        if not row_data.get('name') or str(row_data.get('name', '')).strip() == '':
            continue
        
        try:
            name = str(row_data.get('name', '')).strip()
            location = str(row_data.get('location', '')).strip()
            state = str(row_data.get('state', '') or '').strip()
            description = str(row_data.get('description', '')).strip()
            placement_raw = row_data.get('placement_percentage', 0)
            placement_percentage = float(placement_raw) if placement_raw else 0.0
            rating_raw = row_data.get('rating', 0)
            rating = float(rating_raw) if rating_raw else 0.0

            courses_raw = str(row_data.get('courses', '') or '')
            courses_names = [c.strip() for c in courses_raw.split(',') if c.strip()]
            courses = [
                {
                    "course_name": c_name,
                    "description": "",
                    "duration": "N/A",
                    "fees": 0.0,
                    "category": "Uncategorized"
                } for c_name in courses_names
            ]

            # Tags - comma-separated
            tags_raw = str(row_data.get('tags', '') or '')
            tags = [t.strip() for t in tags_raw.split(',') if t.strip()]

            # Contact details
            contact_details = {}
            email_val = str(row_data.get('email', '') or '').strip()
            phone_val = str(row_data.get('phone', '') or '').strip()
            website_val = str(row_data.get('website', '') or '').strip()
            if email_val:
                contact_details['email'] = email_val
            if phone_val:
                contact_details['phone'] = phone_val
            if website_val:
                contact_details['website'] = website_val

            if not name or not location or not description:
                errors.append(f"Row {row_idx}: Missing required fields (name, location, description)")
                continue

            university = University(
                name=name,
                location=location,
                state=state,
                main_photo="",  # No photo - will show placeholder
                description=description,
                courses=courses,
                placement_percentage=placement_percentage,
                rating=rating,
                tags=tags,
                contact_details=contact_details
            )
            await db.universities.insert_one(university.model_dump())
            created.append(name)

        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
    
    return {
        "message": f"Bulk upload complete. {len(created)} universities created.",
        "created_count": len(created),
        "created": created,
        "errors": errors,
        "error_count": len(errors)
    }

@api_router.get("/universities/{university_id}", response_model=University)
async def get_university(university_id: str):
    university = await db.universities.find_one({"id": university_id}, {"_id": 0})
    if not university:
        raise HTTPException(status_code=404, detail="University not found")
    return university

@api_router.put("/universities/{university_id}", response_model=University)
async def update_university(
    university_id: str,
    university_data: UniversityUpdate,
    current_user: Dict = Depends(get_current_user)
):
    if current_user['role'] == 'manager' and current_user.get('university_id') != university_id:
        raise HTTPException(status_code=403, detail="You can only update your own university")
    
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    update_data = {k: v for k, v in university_data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    result = await db.universities.update_one(
        {"id": university_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="University not found")
    
    university = await db.universities.find_one({"id": university_id}, {"_id": 0})
    return university

@api_router.delete("/universities/{university_id}")
async def delete_university(university_id: str, current_user: Dict = Depends(get_current_user)):
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can delete universities")
    
    result = await db.universities.delete_one({"id": university_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="University not found")
    
    return {"message": "University deleted"}

@api_router.get("/universities/{university_id}/applications/export")
async def export_university_applications(university_id: str, current_user: Dict = Depends(get_current_user)):
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    if current_user['role'] == 'manager' and current_user.get('university_id') != university_id:
        raise HTTPException(status_code=403, detail="You can only export your own university's applications")
    
    university = await db.universities.find_one({"id": university_id}, {"_id": 0})
    if not university:
        raise HTTPException(status_code=404, detail="University not found")
    
    applications = await db.applications.find({"university_id": university_id}, {"_id": 0}).to_list(10000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    
    headers = ['Name', 'Email', 'Phone', 'University Name', 'Course Interest', 'Short Note', 'Date', 'Status']
    ws.append(headers)
    
    header_fill = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for app in applications:
        created_date = datetime.fromisoformat(app['created_at']).strftime('%Y-%m-%d %H:%M')
        ws.append([
            app['name'],
            app['email'],
            app['phone'],
            app['university_name'],
            app['course_interest'],
            app['short_note'],
            created_date,
            app['status']
        ])
    
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    uni_name = university['name'].replace(' ', '_')
    filename = f"{uni_name}_applications.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============ APPLICATION ENDPOINTS ============

@api_router.post("/applications", response_model=ApplicationResponse)
async def create_application(application_data: ApplicationCreate):
    application = Application(**application_data.model_dump())
    await db.applications.insert_one(application.model_dump())
    return ApplicationResponse(**application.model_dump())

@api_router.get("/applications", response_model=List[ApplicationResponse])
async def get_applications(current_user: Dict = Depends(get_current_user)):
    query = {}
    
    if current_user['role'] == 'manager':
        if not current_user.get('university_id'):
            raise HTTPException(status_code=403, detail="No university assigned")
        query['university_id'] = current_user['university_id']
    
    if current_user['role'] == 'student':
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    applications = await db.applications.find(query, {"_id": 0}).to_list(10000)
    return [ApplicationResponse(**app) for app in applications]

@api_router.get("/applications/university/{university_id}", response_model=List[ApplicationResponse])
async def get_applications_by_university(university_id: str, current_user: Dict = Depends(get_current_user)):
    if current_user['role'] == 'manager' and current_user.get('university_id') != university_id:
        raise HTTPException(status_code=403, detail="You can only view your own university's applications")
    
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    applications = await db.applications.find({"university_id": university_id}, {"_id": 0}).to_list(10000)
    return [ApplicationResponse(**app) for app in applications]

@api_router.put("/applications/{application_id}/status")
async def update_application_status(
    application_id: str,
    status: str,
    current_user: Dict = Depends(get_current_user)
):
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    application = await db.applications.find_one({"id": application_id}, {"_id": 0})
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")
    
    if current_user['role'] == 'manager' and application['university_id'] != current_user.get('university_id'):
        raise HTTPException(status_code=403, detail="You can only update your own university's applications")
    
    await db.applications.update_one(
        {"id": application_id},
        {"$set": {"status": status}}
    )
    
    return {"message": "Status updated"}

@api_router.delete("/applications/{application_id}")
async def delete_application(application_id: str, current_user: Dict = Depends(get_current_user)):
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    application = await db.applications.find_one({"id": application_id}, {"_id": 0})
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")
    
    if current_user['role'] == 'manager' and application['university_id'] != current_user.get('university_id'):
        raise HTTPException(status_code=403, detail="You can only delete your own university's applications")
    
    await db.applications.delete_one({"id": application_id})
    
    return {"message": "Application deleted"}

# ============ EXCEL EXPORT ENDPOINT ============

@api_router.get("/applications/export/excel")
async def export_applications_excel(current_user: Dict = Depends(get_current_user)):
    if current_user['role'] not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Unauthorized")
    
    query = {}
    filename = "all_applications.xlsx"
    
    if current_user['role'] == 'manager':
        if not current_user.get('university_id'):
            raise HTTPException(status_code=403, detail="No university assigned")
        query['university_id'] = current_user['university_id']
        
        university = await db.universities.find_one({"id": current_user['university_id']}, {"_id": 0})
        if university:
            uni_name = university['name'].replace(' ', '_')
            filename = f"{uni_name}_applications.xlsx"
    
    applications = await db.applications.find(query, {"_id": 0}).to_list(10000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    
    headers = ['Name', 'Email', 'Phone', 'University Name', 'Course Interest', 'Short Note', 'Date', 'Status']
    ws.append(headers)
    
    header_fill = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for app in applications:
        created_date = datetime.fromisoformat(app['created_at']).strftime('%Y-%m-%d %H:%M')
        ws.append([
            app['name'],
            app['email'],
            app['phone'],
            app['university_name'],
            app['course_interest'],
            app['short_note'],
            created_date,
            app['status']
        ])
    
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============ ADMIN ENDPOINTS ============

@api_router.get("/admin/stats")
async def get_admin_stats(current_user: Dict = Depends(get_current_user)):
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    
    total_universities = await db.universities.count_documents({})
    total_applications = await db.applications.count_documents({})
    total_managers = await db.users.count_documents({"role": "manager"})
    pending_applications = await db.applications.count_documents({"status": "pending"})
    
    return {
        "total_universities": total_universities,
        "total_applications": total_applications,
        "total_managers": total_managers,
        "pending_applications": pending_applications
    }

@api_router.post("/admin/users", response_model=UserResponse)
async def create_user(user_data: UserCreate, current_user: Dict = Depends(get_current_user)):
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    
    existing = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user = User(
        email=user_data.email,
        password_hash=hash_password(user_data.password),
        name=user_data.name,
        role=user_data.role,
        university_id=user_data.university_id
    )
    
    await db.users.insert_one(user.model_dump())
    
    return UserResponse(
        id=user.id,
        email=user.email,
        name=user.name,
        role=user.role,
        university_id=user.university_id
    )

@api_router.get("/admin/users", response_model=List[UserResponse])
async def get_all_users(current_user: Dict = Depends(get_current_user)):
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return [UserResponse(**user) for user in users]

@api_router.put("/admin/users/{user_id}")
async def update_user(
    user_id: str,
    name: Optional[str] = None,
    email: Optional[str] = None,
    password: Optional[str] = None,
    university_id: Optional[str] = None,
    current_user: Dict = Depends(get_current_user)
):
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    
    update_data = {}
    if name:
        update_data['name'] = name
    if email:
        existing = await db.users.find_one({"email": email, "id": {"$ne": user_id}}, {"_id": 0})
        if existing:
            raise HTTPException(status_code=400, detail="Email already in use")
        update_data['email'] = email
    if password:
        update_data['password_hash'] = hash_password(password)
    if university_id is not None:
        update_data['university_id'] = university_id
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    result = await db.users.update_one({"id": user_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "User updated successfully"}

@api_router.delete("/admin/users/{user_id}")
async def delete_user(user_id: str, current_user: Dict = Depends(get_current_user)):
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Admin only")
    
    if user_id == current_user['user_id']:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "User deleted successfully"}

# Include router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_db_client():
    """On startup, create the default admin user if no admin exists yet."""
    existing_admin = await db.users.find_one({"role": "admin"}, {"_id": 0})
    if not existing_admin:
        admin_email    = os.environ.get("ADMIN_EMAIL", "admin@edudham.com")
        admin_password = os.environ.get("ADMIN_PASSWORD", "Admin@123")
        admin_name     = os.environ.get("ADMIN_NAME", "Admin")
        admin = User(
            email=admin_email,
            password_hash=hash_password(admin_password),
            name=admin_name,
            role="admin",
        )
        await db.users.insert_one(admin.model_dump())
        logger.info(f"✅ Default admin created: {admin_email}")
    else:
        logger.info("✅ Admin user already exists — skipping seed.")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()